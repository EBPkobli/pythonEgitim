#pip install openpyxl
#pip install pillow

import openpyxl

#print(openpyxl.__version__)

wb = openpyxl.load_workbook("freelancer.xlsx")

#print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('SheetAdimPythondan')

print(sheet['A1'].value)

sheet['A1'].value = "PYTHON AWESOME!"
wb.save('freelancer.xlsx')

sheet.title = "SheetAdimPythondan"

print(sheet.cell(row=5,column=3).value)

for i in range(1,10):
    print(sheet.cell(row=i,column=3).value)

wb.create_sheet(title="pythonSheet",index=1)
sheet = wb.get_sheet_by_name("pythonSheet")
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 50
wb.save('freelancer.xlsx')

from openpyxl.styles import Font
import random

for i in range(1,30):
    sheet['A'+str(i)].value = random.randint(1,999)
sheet['A5'].font = Font(sz=14,bold=True,italic=True)
wb.save('freelancer.xlsx')